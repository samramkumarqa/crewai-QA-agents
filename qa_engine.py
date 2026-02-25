# qa_engine.py - GEMINI VERSION
import os
os.environ["CREWAI_TELEMETRY_ENABLED"] = "false"
os.environ["OTEL_SDK_DISABLED"] = "true"
os.environ["OPENTELEMETRY_SDK_DISABLED"] = "true"

import sys
import json
import re
import ast
from datetime import datetime

# Import CrewAI components
from crewai import Agent, Crew, Process, Task
from crewai.project import CrewBase, agent, crew, task

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment

# === GEMINI INTEGRATION (FREE) ===
import google.generativeai as genai

load_dotenv()

class GeminiDirectLLM:
    """Direct Google Gemini integration - completely free tier"""
    
    def __init__(self, api_key=None, model="gemini-1.5-flash"):
        """
        Initialize Gemini LLM
        - gemini-1.5-flash: Fast, free tier (60 requests/min)
        - gemini-1.5-pro: More powerful, also has free tier
        """
        self.api_key = api_key or os.getenv("GEMINI_API_KEY")
        if not self.api_key:
            raise ValueError("GEMINI_API_KEY is required. Get one from: https://aistudio.google.com/app/apikey")
        
        # Configure Gemini
        genai.configure(api_key=self.api_key)
        self.model = genai.GenerativeModel(model)
        self.model_name = model
        
        # Print debug info
        print(f"🔑 Gemini API key configured")
        print(f"🤖 Using model: {self.model_name}")
        
        # Test the connection
        try:
            test_response = self.generate("Say 'Gemini is ready!' in one line.")
            print(f"✅ Gemini test successful: {test_response}")
        except Exception as e:
            print(f"⚠️ Gemini test warning: {e}")
    
    def generate(self, prompt, **kwargs):
        """Generate a response from Gemini"""
        try:
            # Set generation config
            generation_config = {
                "temperature": kwargs.get("temperature", 0.0),
                "max_output_tokens": kwargs.get("max_tokens", 1500),
                "top_p": kwargs.get("top_p", 0.8),
                "top_k": kwargs.get("top_k", 40),
            }
            
            # Generate response
            response = self.model.generate_content(
                prompt,
                generation_config=generation_config
            )
            
            return response.text
            
        except Exception as e:
            print(f"❌ Gemini generation error: {str(e)}")
            raise
    
    def __call__(self, messages, **kwargs):
        """Make the LLM callable like CrewAI expects"""
        if isinstance(messages, list):
            # Extract the last user message
            for msg in reversed(messages):
                if isinstance(msg, dict) and msg.get("role") == "user":
                    prompt = msg.get("content", "")
                    break
            else:
                # If no user message found, use the whole list
                prompt = str(messages)
        else:
            prompt = str(messages)
        
        return self.generate(prompt, **kwargs)

# Initialize Gemini LLM
try:
    llm = GeminiDirectLLM()
    print("✅ Google Gemini integration successful!")
except Exception as e:
    print(f"❌ Failed to initialize Gemini: {str(e)}")
    print("\n📝 To fix this:")
    print("1. Go to https://aistudio.google.com/app/apikey")
    print("2. Click 'Create API Key'")
    print("3. Copy the key")
    print("4. Add to Streamlit secrets: GEMINI_API_KEY = 'your-key-here'")
    print("5. Or create .env file with: GEMINI_API_KEY=your-key-here")
    raise
# ======================================

# ---------- Helpers ----------
def parse_list_of_dicts(text):
    if isinstance(text, list):
        return text
    if not isinstance(text, str):
        return []

    try:
        return ast.literal_eval(text)
    except:
        return []

def normalize_edge(e):
    if isinstance(e, dict):
        return str(e.get("description", e))
    return str(e)

def format_steps(steps):
    if isinstance(steps, list):
        return "\n".join([f"{i+1}. {s}" for i, s in enumerate(steps)])
    return str(steps)

def normalize_steps(raw_steps):
    if not raw_steps:
        return ""

    if isinstance(raw_steps, list):
        return "\n".join([f"{i+1}. {s.get('step', s) if isinstance(s, dict) else s}"
                          for i, s in enumerate(raw_steps)])

    if isinstance(raw_steps, dict):
        return f"1. {raw_steps.get('step', raw_steps)}"

    if isinstance(raw_steps, str):
        parts = re.split(r'(?=Enter |Click |Verify |Select |Login |Log in |Open |Submit )', raw_steps)
        parts = [p.strip() for p in parts if p.strip()]
        if len(parts) > 1:
            return "\n".join([f"{i+1}. {p}" for i, p in enumerate(parts)])
        return raw_steps

    return str(raw_steps)

def normalize_list(data):
    if not data: return []
    if isinstance(data, (dict, str)): return [data]
    if isinstance(data, list): return data
    return [str(data)]

def safe_json(text):
    try:
        return json.loads(text)
    except:
        m = re.search(r"\[.*\]", text, re.DOTALL)
        if m:
            try:
                return json.loads(m.group())
            except:
                return []
        return []

# ---------- Excel ----------
def export_excel(brd, scenarios, tcs, edges, auto):
    brd = normalize_list(brd)
    scenarios = normalize_list(scenarios)
    tcs = normalize_list(tcs)
    edges = normalize_list(edges)
    auto = normalize_list(auto)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "BRD Analysis"
    ws1.append(["Module", "Description"])
    brd_rows = []
    for r in brd:
        if isinstance(r, str) and r.strip().startswith("["):
            brd_rows.extend(parse_list_of_dicts(r))
        elif isinstance(r, dict):
            brd_rows.append(r)

    for r in brd_rows:
        ws1.append([r.get("module",""), r.get("description","")])

    ws2 = wb.create_sheet("Test Scenarios")
    ws2.append(["ID", "Description"])
    for s in scenarios:
        ws2.append([s.get("id",""), s.get("description","")]) if isinstance(s, dict) else ws2.append(["", str(s)])

    ws3 = wb.create_sheet("Detailed Test Cases")
    ws3.append(["ID", "Scenario", "Steps", "Expected Result", "Type"])
    for t in tcs:
        if isinstance(t, dict):
            ws3.append([
                t.get("id",""),
                t.get("scenario",""),
                normalize_steps(t.get("steps")),
                t.get("expected_result",""),
                t.get("test_type","")
            ])
        else:
            ws3.append(["", str(t), "", "", ""])

    ws4 = wb.create_sheet("Edge Cases")
    ws4.append(["ID", "Scenario", "Steps", "Expected Result"])

    edge_rows = []
    for e in edges:
        if isinstance(e, str) and e.strip().startswith("["):
            edge_rows.extend(parse_list_of_dicts(e))
        elif isinstance(e, dict):
            edge_rows.append(e)

    for e in edge_rows:
        ws4.append([
            e.get("id",""),
            e.get("scenario",""),
            normalize_steps(e.get("steps")),
            e.get("expected_result","")
        ])

    ws5 = wb.create_sheet("Automation Candidates")
    ws5.append(["ID", "Reason"])
    for a in auto:
        ws5.append([a.get("id",""), a.get("reason","")]) if isinstance(a, dict) else ws5.append(["", str(a)])

    for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    name = f"QA_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(name)
    return name

# ---------- Crew ----------
@CrewBase
class QACrew():
    agents_config = "config/agents.yaml"
    tasks_config = "config/tasks.yaml"

    @agent
    def lead_qa(self): 
        return Agent(
            config=self.agents_config["lead_qa"], 
            llm=llm,
            verbose=True,
            allow_delegation=False
        )

    @agent
    def scenario_designer(self): 
        return Agent(
            config=self.agents_config["scenario_designer"], 
            llm=llm,
            verbose=True,
            allow_delegation=False
        )

    @agent
    def testcase_writer(self): 
        return Agent(
            config=self.agents_config["testcase_writer"], 
            llm=llm,
            verbose=True,
            allow_delegation=False
        )

    @agent
    def qa_reviewer(self): 
        return Agent(
            config=self.agents_config["qa_reviewer"], 
            llm=llm,
            verbose=True,
            allow_delegation=False
        )

    @task
    def brd_analysis(self): 
        return Task(config=self.tasks_config["brd_analysis"], agent=self.lead_qa())

    @task
    def test_scenarios(self): 
        return Task(config=self.tasks_config["test_scenarios"], agent=self.scenario_designer())

    @task
    def detailed_testcases(self): 
        return Task(config=self.tasks_config["detailed_testcases"], agent=self.testcase_writer())

    @task
    def edge_case_review(self): 
        return Task(config=self.tasks_config["edge_case_review"], agent=self.qa_reviewer())

    @task
    def automation_candidates(self): 
        return Task(config=self.tasks_config["automation_candidates"], agent=self.qa_reviewer())

    @crew
    def qacrew(self):
        return Crew(
            agents=self.agents, 
            tasks=self.tasks, 
            process=Process.sequential, 
            verbose=True,
            memory=False,
            cache=False
        )